import json
import os
import argparse
import pathlib

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from utils import get_account_id, insert_candidate, get_all_vacancies, connect_candidate_to_vacancy, get_vacancy, \
    get_all_statuses, get_status, get_fill_quota


def main():
    parser = argparse.ArgumentParser(description='Process arguments to access API')
    parser.add_argument(
        '-t',
        '--token',
        help='Huntflow api token',
        default=None,
        type=str
    )
    parser.add_argument(
        '-p',
        '--path',
        help='Huntflow database path',
        default=None,
        type=pathlib.Path
    )
    args = parser.parse_args()
    token = args.token
    path = args.path

    if token is None:
        token = input('Enter Huntflow api token: ')

    if path is None:
        path = pathlib.Path(input('Enter Huntflow database path: '))

    xlsx_instance = openpyxl.load_workbook(path.resolve(strict=True))

    work_sheet: Worksheet = xlsx_instance[xlsx_instance.sheetnames[0]]

    data_dump = {
        'start': 1
    }
    if os.path.exists('dump.json'):
        with open('dump.json') as f:
            data_dump = json.load(f)

    BASE_API_URL = "https://dev-100-api.huntflow.dev/"

    account_id = get_account_id(token, BASE_API_URL)

    all_vacancies = get_all_vacancies(token, BASE_API_URL, account_id)
    all_statuses = get_all_statuses(token, BASE_API_URL, account_id)

    for row in work_sheet.iter_rows(min_row=data_dump['start'] + 1):
        try:
            data_dump['start'] += 1
            candidate_cv = insert_candidate(token, BASE_API_URL, account_id, row)
            vacancy = get_vacancy(all_vacancies, candidate_cv['position'])
            status = get_status(all_statuses, row[4].value.strip())
            connect_candidate_to_vacancy(token,
                                         BASE_API_URL,
                                         account_id,
                                         row,
                                         vacancy['id'],
                                         status['id'],
                                         candidate_cv)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            with open('dump.json', 'w') as f:
                json.dump(data_dump, f)
            break
    else:
        if os.path.exists('dump.json'):
            os.remove('dump.json')




if __name__ == '__main__':
    main()
